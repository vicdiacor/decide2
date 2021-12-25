from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework.test import APITestCase

from http import HTTPStatus

from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token

from base import mods
from base.tests import BaseTestCase, SeleniumBaseTestCase
from selenium.webdriver.common.by import By

from django.contrib import auth
from django.contrib.auth.models import User, Group
from selenium.webdriver.support.ui import Select
import os
import time
import openpyxl

from selenium.common.exceptions import NoSuchElementException

from authentication.import_and_export import *       



class AuthTestCase(APITestCase):

    def setUp(self):
        self.client = APIClient()
        mods.mock_query(self.client)
        u = User(username='voter1')
        u.set_password('123')
        u.save()

        u2 = User(username='admin')
        u2.set_password('admin')
        u2.is_superuser = True
        u2.save()

    def tearDown(self):
        self.client = None

    def test_login(self):
        data = {'username': 'voter1', 'password': '123'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)

        token = response.json()
        self.assertTrue(token.get('token'))

    def test_login_fail(self):
        data = {'username': 'voter1', 'password': '321'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_getuser(self):
        data = {'username': 'voter1', 'password': '123'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.json()

        response = self.client.post('/authentication/getuser/', token, format='json')
        self.assertEqual(response.status_code, 200)

        user = response.json()
        self.assertEqual(user['id'], 1)
        self.assertEqual(user['username'], 'voter1')

    def test_getuser_invented_token(self):
        token = {'token': 'invented'}
        response = self.client.post('/authentication/getuser/', token, format='json')
        self.assertEqual(response.status_code, 404)

    def test_getuser_invalid_token(self):
        data = {'username': 'voter1', 'password': '123'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Token.objects.filter(user__username='voter1').count(), 1)

        token = response.json()
        self.assertTrue(token.get('token'))

        response = self.client.post('/authentication/logout/', token, format='json')
        self.assertEqual(response.status_code, 200)

        response = self.client.post('/authentication/getuser/', token, format='json')
        self.assertEqual(response.status_code, 404)

    def test_logout(self):
        data = {'username': 'voter1', 'password': '123'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Token.objects.filter(user__username='voter1').count(), 1)

        token = response.json()
        self.assertTrue(token.get('token'))

        response = self.client.post('/authentication/logout/', token, format='json')
        self.assertEqual(response.status_code, 200)

        self.assertEqual(Token.objects.filter(user__username='voter1').count(), 0)

    def test_register_bad_permissions(self):
        data = {'username': 'voter1', 'password': '123'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.json()

        token.update({'username': 'user1'})
        response = self.client.post('/authentication/register/', token, format='json')
        self.assertEqual(response.status_code, 401)

    def test_register_bad_request(self):
        data = {'username': 'admin', 'password': 'admin'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.json()

        token.update({'username': 'user1'})
        response = self.client.post('/authentication/register/', token, format='json')
        self.assertEqual(response.status_code, 400)

    def test_register_user_already_exist(self):
        data = {'username': 'admin', 'password': 'admin'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.json()

        token.update(data)
        response = self.client.post('/authentication/register/', token, format='json')
        self.assertEqual(response.status_code, 400)

    def test_register(self):
        data = {'username': 'admin', 'password': 'admin'}
        response = self.client.post('/authentication/login/', data, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.json()

        token.update({'username': 'user1', 'password': 'pwd1'})
        response = self.client.post('/authentication/register/', token, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            sorted(list(response.json().keys())),
            ['token', 'user_pk']
        )

class SeleniumTestCase(SeleniumBaseTestCase):

    def test_simpleCorrectLogin(self):
        self.login()
        #In case of a correct loging, a element with id 'user-tools' is shown in the upper right part
        self.assertTrue(len(self.driver.find_elements_by_id('user-tools'))==1)

    def test_incorrect_login(self):        
        self.login('notanuser', 'notapassword')

        self.assertFalse(len(self.driver.find_elements_by_id('user-tools'))==1)
    
    def test_selenium_extension(self):
        self.driver.get(f"{self.live_server_url}/admin/login/?next=/admin/")
        self.driver.set_window_size(909, 1016)
        self.driver.find_element(By.ID, "id_username").send_keys("admin")
        self.driver.find_element(By.ID, "id_password").send_keys("qwerty")
        self.driver.find_element(By.CSS_SELECTOR, ".submit-row > input").click()
        
        self.assertEquals(self.driver.current_url, f'{self.live_server_url}/admin/')
        self.driver.close()

        
    def test_selenium_correct_registration_and_login(self):
        self.driver.get(f"{self.live_server_url}/authentication/registrarse/")
        self.driver.find_element(By.ID, "id_username").send_keys("testuser")
        self.driver.find_element(By.ID, "id_first_name").send_keys("test")
        self.driver.find_element(By.ID, "id_last_name").send_keys("prueba")
        self.driver.find_element(By.ID, "id_email").send_keys("prueba@gmail.com")
        self.driver.find_element(By.ID, "id_password1").send_keys("decide1234")
        self.driver.find_element(By.ID, "id_password2").send_keys("decide1234")
        self.driver.find_element_by_xpath('//button["Registrarse"]').click()
        
        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
        self.driver.find_element(By.XPATH, '//a["Cerrar sesión"]')
        user = auth.get_user(self.client)
        self.assertFalse(user.is_authenticated)
        
        self.driver.get(f"{self.live_server_url}/authentication/iniciar_sesion/")
        self.driver.find_element(By.ID, "id_username").send_keys("testuser")
        self.driver.find_element(By.ID, "id_password").send_keys("decide1234")
        self.driver.find_element_by_xpath('//button["Inicie sesión"]').click()
        
        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
        
    def test_selenium_different_passwords(self):
        self.driver.get(f"{self.live_server_url}/authentication/registrarse/")
        self.driver.find_element(By.ID, "id_username").send_keys("testuser")
        self.driver.find_element(By.ID, "id_first_name").send_keys("test")
        self.driver.find_element(By.ID, "id_last_name").send_keys("prueba")
        self.driver.find_element(By.ID, "id_email").send_keys("prueba@gmail.com")
        self.driver.find_element(By.ID, "id_password1").send_keys("decide1234")
        self.driver.find_element(By.ID, "id_password2").send_keys("decide1235")
        self.driver.find_element_by_xpath('//button["Registrarse"]').click()
        
        self.driver.find_element_by_xpath('//p["Los dos campos de contraseña no coinciden."]')
        
    def test_selenium_username_already_exists(self):
        self.driver.get(f"{self.live_server_url}/authentication/registrarse/")
        self.driver.find_element(By.ID, "id_username").send_keys("testuser")
        self.driver.find_element(By.ID, "id_first_name").send_keys("test")
        self.driver.find_element(By.ID, "id_last_name").send_keys("prueba")
        self.driver.find_element(By.ID, "id_email").send_keys("prueba@gmail.com")
        self.driver.find_element(By.ID, "id_password1").send_keys("decide1234")
        self.driver.find_element(By.ID, "id_password2").send_keys("decide1234")
        self.driver.find_element_by_xpath('//button["Registrarse"]').click()
        
        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
        self.driver.find_element(By.XPATH, '//a["Cerrar sesión"]')
        user = auth.get_user(self.client)
        self.assertFalse(user.is_authenticated)
        
        self.driver.get(f"{self.live_server_url}/authentication/registrarse/")
        self.driver.find_element(By.ID, "id_username").send_keys("testuser")
        self.driver.find_element(By.ID, "id_first_name").send_keys("test")
        self.driver.find_element(By.ID, "id_last_name").send_keys("prueba")
        self.driver.find_element(By.ID, "id_email").send_keys("prueba1@gmail.com")
        self.driver.find_element(By.ID, "id_password1").send_keys("decide1234")
        self.driver.find_element(By.ID, "id_password2").send_keys("decide1234")
        self.driver.find_element_by_xpath('//button["Registrarse"]').click()
        
        self.driver.find_element_by_xpath('//p["Ya existe un usuario con este nombre."]')  
        
class IniciarSesionTestCase(TestCase):
    
    def setUp(self):
        
        self.credentials = {
            'username': 'testuser',
            'password': 'decide1234'}
        User.objects.create_user(**self.credentials)
        
    def tearDown(self):
        self.client = None
        
    def test_get_form(self):
        response = self.client.get("/authentication/iniciar_sesion/")

        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Inicie sesión en DECIDE</h2>", html=True)
    
    def test_post_form(self):
        response = self.client.post("/authentication/iniciar_sesion/", data={"username": "testuser", "password": "decide1234"})
        
        self.assertEqual(response.status_code, HTTPStatus.FOUND)
        self.assertEqual(response["Location"], "/")

        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
    def test_post_incorrect_form(self):
        #Mal usuario
        response = self.client.post("/authentication/iniciar_sesion/", data={"username": "testuser_error", "password": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "Por favor, introduzca un nombre de usuario y clave correctos. Observe que ambos campos pueden ser sensibles a mayúsculas.", html=True)
        
        #Mala contraseña
        response = self.client.post("/authentication/iniciar_sesion/", data={"username": "testuser_error", "password": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "Por favor, introduzca un nombre de usuario y clave correctos. Observe que ambos campos pueden ser sensibles a mayúsculas.", html=True)
        
        #Campos vacíos
        response = self.client.post("/authentication/iniciar_sesion/", data={"username": "", "password": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Inicie sesión en DECIDE</h2>", html=True)
        

class CerrarSesionTestCase(TestCase):
    
    def setUp(self):
        
        self.credentials = {
            'username': 'testuser',
            'password': 'decide1234'}
        User.objects.create_user(**self.credentials)
        
    def tearDown(self):
        self.client = None
        
    def test_cerrar_sesion(self):
        
        response = self.client.post("/authentication/iniciar_sesion/", data={"username": "testuser", "password": "decide1234"})
        
        self.assertEqual(response.status_code, HTTPStatus.FOUND)
        self.assertEqual(response["Location"], "/")

        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)

        response = self.client.get("/authentication/cerrar_sesion/")

        self.assertEqual(response.status_code, HTTPStatus.OK)
        
        user = auth.get_user(self.client)
        self.assertFalse(user.is_authenticated)        
        
class RegistrarUsuarioTestCase(TestCase):
    
    def tearDown(self):
        self.client = None
        
    def test_good_registration(self):
        username_t="testuser"
        password_t="decide1234"
        try:
            user = User.objects.get(username="testuser")
            exists=True
        except User.DoesNotExist:
            exists=False
        
        self.assertFalse(exists)       
        
        response=self.client.get("/authentication/registrarse/")
        self.assertEqual(response.status_code, HTTPStatus.OK)
        
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "testuser@example.com","password1": "decide1234", "password2": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.FOUND)
        self.assertEqual(response["Location"], "/")
        
        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
    def test_wrong_registration(self):
        
        username_t="testuser"
        password_t="decide1234"
        try:
            user = User.objects.get(username="testuser")
            exists=True
        except User.DoesNotExist:
            exists=False
        
        self.assertFalse(exists)          
        
        #Campos vacíos
        response=self.client.post("/authentication/registrarse/", data={"username": "", "first_name": "", "last_name": "","email": "","password1": "", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        
        #Solo username
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "", "last_name": "","email": "","password1": "", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Solo username y first_name
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "","email": "","password1": "", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Solo username, first_name y last_name
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "","password1": "", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
       
        #Solo username, first_name, last_name y email
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "","password1": "", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Sin password2
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "prueba@gmail.com","password1": "decide1234", "password2": ""})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Sin password2 --> Mirar en internet como comprobar error de form
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "prueba@gmail.com","password1": "decide1234", "password2": "decide1235"})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Mal email
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "prueba","password1": "decide1234", "password2": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)
        
        #Usuario ya registrado --> Mirar en internet como comprobar error de form
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "testuser@example.com","password1": "decide1234", "password2": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.FOUND)
        self.assertEqual(response["Location"], "/")
        
        user = User.objects.get(username='testuser')
        self.assertTrue(user.is_authenticated)
        
        response=self.client.post("/authentication/registrarse/", data={"username": "testuser", "first_name": "testuser", "last_name": "testuser","email": "prueba@gmail.com","password1": "decide1234", "password2": "decide1234"})
        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertContains(response, "<h2>Formulario de registro en DECIDE</h2>", html=True)




class ImportAndExportGroupTestCase(TestCase):

    def setUp(self):
        g1 = Group(name='Grupo 1', pk=100)
        g1.save()

        g2 = Group(name='Grupo 2', pk=101)
        g2.save()

        u1 = User(username='username1', password='password')
        u1.save()
        u1.groups.set([g1, g2])
        u1.save()

        u2 = User(username='username2', password='password')
        u2.save()
        u2.groups.set([g1])
        u2.save()

        u3 = User(username='username3', password='password')
        u3.save()
        u3.groups.set([g1])
        u3.save()

        u4 = User(username='username4', password='password')
        u4.save()
        u4.groups.set([g1])
        u4.save()


        u5 = User(username='username5', password='password')
        u5.save()
        u5.groups.set([g1])
        u5.save()

        return super().setUp()

        

    # Prueba la función "readTxtFile"
    def test_read_txt_file(self):

        file1 = open('authentication/files/testfiles/testgroup1.txt', 'rb')
        users_list_1 = readTxtFile(file1)

        # Comproba cada usuario de la lista devuelta
        self.assertTrue(len(users_list_1)==5)
        self.assertEquals(users_list_1[0], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username5'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))
        self.assertEquals(users_list_1[3], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[4], User.objects.get(username='username4'))

        # Comprueba que devuelve None si algún usuario del fichero no existe
        file2 = open('authentication/files/testfiles/testgroup2.txt', 'rb')
        users_list_2 = readTxtFile(file2)
        self.assertEquals(users_list_2, None)


    # Prueba la función "createGroup"
    def test_create_group(self):

        # Comprueba que crea grupo nuevo
        name_1 = 'Grupo 3'
        users_list_1 = User.objects.all()
        res1 = createGroup(name_1, users_list_1)
        self.assertTrue(res1)
        
        res2 = True
        try:
            Group.objects.get(name='Grupo 3')
        except:
            res2=False

        self.assertTrue(res2)


        # Comprueba que actualiza 'Grupo 1'
        name_2 = 'Grupo 1'
        users_list_2 = []
        u1 = User.objects.get(username='username1')
        users_list_2.append(u1)


        res3 = createGroup(name_2, users_list_2)
        self.assertFalse(res3)
        self.assertEquals(len(User.objects.filter(groups__name='Grupo 1')), 1)


    # Prueba la función "readExcelFile"
    def test_read_excel_file(self):
        users_list_1 = readExcelFile('authentication/files/testfiles/testgroup1.xlsx')

        # Comproba cada usuario de la lista devuelta
        self.assertTrue(len(users_list_1)==5)
        self.assertEquals(users_list_1[0], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username5'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))
        self.assertEquals(users_list_1[3], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[4], User.objects.get(username='username4'))

        # Comprueba que devuelve None si algún usuario del fichero no existe
        users_list_2 = readExcelFile('authentication/files/testfiles/testgroup2.xlsx')
        self.assertEquals(users_list_2, None)


    # Prueba la función "auxUsersList"
    def test_aux_users_list(self):
        username_list_1 = ['username1', 'username2', 'username3']
        users_list_1 = auxUsersList(username_list_1)
        
        self.assertEquals(users_list_1[0], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))

        username_list_2 = ['username1', 'username2', 'usernamequenoexiste']
        users_list_2 = auxUsersList(username_list_2)

        self.assertEquals(users_list_2, None)


    # Prueba la función "writeInExcelUsernames"
    def test_write_in_excel_usernames(self):
        # Fichero correcto
        users = User.objects.all()
        path = 'authentication/files/temp_export.xlsx'
        name = 'temp_export.xlsx'
        writeInExcelUsernames(users, path, name)

        # Leer el fichero excel
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        max_row = sheet.max_row
        username_list = []
        for i in range(1, max_row+ 1):
            cell = sheet.cell(row = i, column = 1)
            username_list.append(cell.value)

        # Comprueba cada usuario
        for i in range(0, len(users)):
            self.assertEquals(users[i].username, username_list[i])


class ImportAndExportGroupSeleniumTestCase(SeleniumBaseTestCase):

    def setUp(self):
        g1 = Group(name='Grupo 1', pk=100)
        g1.save()

        g2 = Group(name='Grupo 2', pk=101)
        g2.save()

        u1 = User(username='username1', password='password')
        u1.save()
        u1.groups.set([g1, g2])
        u1.save()

        u2 = User(username='username2', password='password')
        u2.save()
        u2.groups.set([g1])
        u2.save()

        u3 = User(username='username3', password='password')
        u3.save()
        u3.groups.set([g1])
        u3.save()

        u4 = User(username='username4', password='password')
        u4.save()
        u4.groups.set([g1])
        u4.save()


        u5 = User(username='username5', password='password')
        u5.save()
        u5.groups.set([g1])
        u5.save()

        return super().setUp()


    def test_import_group(self):
        self.driver.get(f"{self.live_server_url}/authentication/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 3')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/authentication/files/testfiles/testgroup1.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        self.driver.find_element_by_class_name('success')

        self.driver.find_element_by_id('id_name').clear()
        self.driver.find_element_by_id('id_name').send_keys('Grupo 4')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/authentication/files/testfiles/testgroup1.xlsx")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        self.driver.find_element_by_class_name('success')

        # Comprueba que existen 4 grupos (setUp + 2 grupos creados)
        self.login()
        self.driver.get(f"{self.live_server_url}/admin/auth/group/")
        self.assertEquals(len(self.driver.find_elements_by_class_name(name='field-__str__')), 4)


    # Prueba si se introduce un nombre de grupo ya existente
    def test_import_group_name_already_exists(self):
        self.driver.get(f"{self.live_server_url}/authentication/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 2')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/authentication/files/testfiles/testgroup3.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        # Comprueba que aparece un mensaje de éxito
        self.driver.find_element_by_class_name('success')


        # Comprueba que username1 no pertenece al grupo 2
        self.login()
        self.driver.get(f"{self.live_server_url}/admin/auth/user/")
        self.driver.find_element_by_link_text("username1").click()
        result = True
        try:
            self.driver.find_element_by_xpath("//select[@name='groups']/option[text()='Grupo 2']")
        except NoSuchElementException:
            result = False

        self.assertFalse(result)


        # Comprueba que username2 sí pertenece al grupo 2
        self.driver.get(f"{self.live_server_url}/admin/auth/user/")
        self.driver.find_element_by_link_text("username2").click()
        result = True
        try:
            self.driver.find_element_by_xpath("//select[@name='groups']/option[text()='Grupo 2']")
        except NoSuchElementException:
            result = False

        self.assertTrue(result)



    # Prueba si se introduce un fichero con un username que no existe
    def test_import_wrong_username(self):
        self.driver.get(f"{self.live_server_url}/authentication/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 2')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/authentication/files/testfiles/testgroup2.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        # Comprueba que aparece un mensaje de error
        self.driver.find_element_by_class_name('error')



    def test_export_group(self):
        self.driver.get(f"{self.live_server_url}/authentication/groups/export/")
        select = Select(self.driver.find_element_by_id('id_group'))
        select.select_by_visible_text('Grupo 1')
        self.driver.find_element_by_xpath("//input[@value='Exportar']").click()

        # Espero a que termine la descarga
        time.sleep(3)
        # Obtenemos la ruta de descarga
        #download_file_path = os.path.join(os.path.expanduser('~'), 'Descargas/export_group.xlsx') 
        # Compruebo si el fichero existe
        self.assertEquals(os.path.exists('export_group.xlsx'), True)

        # Compruebo si todos los usuarios del grupo están
        workbook = openpyxl.load_workbook('export_group.xlsx')
        sheet = workbook.active

        # Lee el excel y almacena en username_list los nombres de usuario
        usernames = ['username1', 'username2', 'username3', 'username4', 'username5']
        max_row = sheet.max_row
        for i in range(1, max_row+ 1):
            cell = sheet.cell(row = i, column = 1)
            self.assertEquals(cell.value in usernames, True)

        # Eliminamos el fichero descargado
        os.remove("export_group.xlsx") 
