import time
import random
from django.contrib.auth.models import User, UserManager, Group
from django.test import TestCase, Client
from base.tests import SeleniumBaseTestCase
from rest_framework.test import APIClient

from .models import Census, ParentGroup
from base import mods
from base.tests import BaseTestCase
import logging as log


from django.test import TestCase
from rest_framework.test import APIClient

from django.contrib.auth.models import User

from base import mods
from base.tests import BaseTestCase, SeleniumBaseTestCase

from django.contrib.auth.models import User, Group
from selenium.webdriver.support.ui import Select
import os
import time
import openpyxl

from selenium.common.exceptions import NoSuchElementException

from census.import_and_export import * 

import re


class CensusTestCase(BaseTestCase):

    def setUp(self):
        super().setUp()
        self.census = Census(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        response = self.client.get(
            '/census/{}/?voter_id={}'.format(1, 2), format='json')
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), 'Invalid voter')

        response = self.client.get(
            '/census/{}/?voter_id={}'.format(1, 1), format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), 'Valid voter')

    def test_list_voting(self):
        response = self.client.get(
            '/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.get(
            '/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get(
            '/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'voters': [1]})

    def test_add_new_voters_conflict(self):
        data = {'voting_id': 1, 'voters': [1]}
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {'voting_id': 2, 'voters': [1, 2, 3, 4]}
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(data.get('voters')), Census.objects.count() - 1)

    def test_destroy_voter(self):
        data = {'voters': [1]}
        response = self.client.delete(
            '/census/{}/'.format(1), data, format='json')
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, Census.objects.count())


class ParentGroupTestCase(BaseTestCase):

    def setUp(self):
        super().setUp()
        self.group = ParentGroup(name='test_group', isPublic=True)
        self.group.save()

    def tearDown(self):
        super().tearDown()
        self.group = None

    def test_list_groups(self):
        response = self.client.get('/admin/census/parentgroup/')
        self.assertEqual(response.status_code, 302)

    def test_create_group(self):

        self.login()
        data = {'name': 'test_group', 'isPublic': True}
        response = self.client.post('/admin/census/parentgroup/add/', data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual('test_group', ParentGroup.objects.get(
            name='test_group').name)

    # def test_delete_group(self):

    #     self.login()
    #     response = self.client.post('/admin/census/parentgroup/{}/delete/'.format(self.group.id))
    #     print(response)
    #     self.assertEqual(response.status_code, 302)
    #     self.assertEqual(0, ParentGroup.objects.count())


class GroupOperationsAPITestCases(BaseTestCase):
    UNION_URL = '/census/union'
    INTERSECTION_URL = '/census/intersection'
    DIFFERENCE_URL = '/census/difference'

    def setUp(self):
        super().setUp()

        user1 = User(username='user1')
        user1.set_password('user1')
        user1.save()

        user2 = User(username='user2')
        user2.set_password('user2')
        user2.save()

        user3 = User(username='user3')
        user3.set_password('user3')
        user3.save()

        user4 = User(username='user4')
        user4.set_password('user4')
        user4.save()

        group1 = ParentGroup.objects.create(name='group1')
        group1.voters.set([user1, user2, user3])

        group2 = ParentGroup.objects.create(name='group2')
        group2.voters.set([user1, user2, user4])

        group3 = ParentGroup.objects.create(name='group3')

        self.groups = [group1, group2]
        self.users = [user1, user2, user3, user4]

    def tearDown(self):
        super().tearDown()

    def test_group_operation_wrong_name(self):
        data = {'base_group': 'group1', 'groups': [
            'group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(self.UNION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': '', 'base_group': 'group1',
                'groups': ['group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 2, 'base_group': 'group1',
                'groups': ['group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(self.DIFFERENCE_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': '  ', 'base_group': 'group1', 'groups': [
            'group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(self.UNION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_group_operation_name_already_exist(self):
        data = {'name': 'group1', 'base_group': 'group1',
                'groups': ['group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(self.UNION_URL, data, format='json')
        self.assertEqual(response.status_code, 409)

    def test_group_operation_base_groups(self):
        data = {'name': 'test', 'groups': ['group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 2,
                'groups': ['group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_group_operation_groups(self):
        data = {'name': 'test', 'base_group': 'group1', 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 'group1',
                'groups': 'group2', 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 'group1',
                'groups': ['group2', 2], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 'group1',
                'groups': [], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_group_operation_wrong_public(self):
        data = {'name': 'test', 'base_group': 'group1', 'groups': ['group2']}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 'group1', 'groups': [
            'group2'], 'is_public': 'True'}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_group_operation_without_being_member(self):
        data = {'name': 'test', 'base_group': 'group3', 'groups': [
            'group1', 'group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 401)

        data = {'name': 'test', 'base_group': 'group1', 'groups': [
            'group3', 'group2'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 401)

        data = {'name': 'test', 'base_group': 'ula', 'groups': [
            'group2', 'group1'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

        data = {'name': 'test', 'base_group': 'group1', 'groups': [
            'group2', 'ula'], 'is_public': True}

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 400)

    def test_group_union(self):
        data = {'name': 'union', 'base_group': 'group1', 'groups': [
            'group2'], 'is_public': True}

        response = self.client.post(self.UNION_URL, data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='user1', password='user1')
        response = self.client.post(self.UNION_URL, data, format='json')
        self.assertEqual(response.status_code, 201)

        union = ParentGroup.objects.get(name='union')
        self.assertEqual(len(union.voters.all()), len(self.users))

    def test_group_intersection(self):
        data = {'name': 'intersection', 'base_group': 'group1',
                'groups': ['group2'], 'is_public': False}

        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='user1', password='user1')
        response = self.client.post(
            self.INTERSECTION_URL, data, format='json')
        self.assertEqual(response.status_code, 201)

        intersection = ParentGroup.objects.get(name='intersection')
        self.assertEqual(len(intersection.voters.all()), 2)

    def test_group_difference(self):
        data = {'name': 'difference', 'base_group': 'group1', 'groups': [
            'group2'], 'is_public': True}

        response = self.client.post(self.DIFFERENCE_URL, data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='user1', password='user1')
        response = self.client.post(self.DIFFERENCE_URL, data, format='json')
        self.assertEqual(response.status_code, 201)

        difference = ParentGroup.objects.get(name='difference')
        self.assertEquals(len(difference.voters.all()), 1)



class GroupOperationsTestCases(SeleniumBaseTestCase):
    url = '/census/operations/'

    def setUp(self):
        super().setUp()

        user1 = User(username='user1')
        user1.set_password('user1')
        user1.save()

        user2 = User(username='user2')
        user2.set_password('user2')
        user2.save()

        user3 = User(username='user3')
        user3.set_password('user3')
        user3.save()

        user4 = User(username='user4')
        user4.set_password('user4')
        user4.save()

        group1 = ParentGroup.objects.create(name='group1')
        group1.voters.set([user1, user2, user3])

        group2 = ParentGroup.objects.create(name='group2')
        group2.voters.set([user1, user2, user4])

        group3 = ParentGroup.objects.create(name='group3')

        self.groups = [group1, group2]
        self.users = [user1, user2, user3, user4]

    def tearDown(self):
        super().tearDown()

    def test_get_form(self):
        self.driver.get(self.live_server_url + self.url)
        self.assertEquals(self.driver.current_url,
                          self.live_server_url + '/authentication/iniciar_sesion/?next=/census/operations/')

        self.auth_login(username='user1', password='user1')
        self.driver.get(f'{self.live_server_url + self.url}')

        self.assertTrue(
            len(self.driver.find_elements_by_css_selector('body > form')) == 1
        )

    def test_group_post(self):
        self.auth_login(username='user1', password='user1')
        self.driver.get(self.live_server_url + self.url)

        self.driver.find_element_by_css_selector(
            '#id_group_name').send_keys('union')

        self.driver.find_element_by_css_selector(
            '#id_base_group > option:nth-child(2)').click()

        self.driver.find_element_by_css_selector(
            '#id_groups > option:nth-child(2)').click()

        self.driver.find_element_by_css_selector('#id_is_public').click()

        self.driver.find_element_by_css_selector(
            '#id_operation > option:nth-child(1)').click()

        self.driver.find_element_by_css_selector(
            'body > form > input[type=submit]:nth-child(3)').click()

        self.assertEquals(ParentGroup.objects.filter(name='union').count(), 1)



class ImportAndExportGroupTestCase(TestCase):

    def setUp(self):
        g1 = ParentGroup(name='Grupo 1', pk=100)
        g1.save()

        g2 = ParentGroup(name='Grupo 2', pk=101)
        g2.save()

        u1 = User(username='username1', password='password')
        u1.save()
        u1.groups.set([g1, g2])
        u1.save()

        u2 = User(username='username2', password='password')
        u2.save()
        u2.groups.set([g1])
        u2.save()

        u3 = User(username='username3', password='password')
        u3.save()
        u3.groups.set([g1])
        u3.save()

        u4 = User(username='username4', password='password')
        u4.save()
        u4.groups.set([g1])
        u4.save()


        u5 = User(username='username5', password='password')
        u5.save()
        u5.groups.set([g1])
        u5.save()

        return super().setUp()

        

    # Prueba la función "readTxtFile"
    def test_read_txt_file(self):

        file1 = open('census/files/testfiles/testgroup1.txt', 'rb')
        users_list_1 = readTxtFile(file1)

        # Comproba cada usuario de la lista devuelta
        self.assertTrue(len(users_list_1)==5)
        self.assertEquals(users_list_1[0], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username5'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))
        self.assertEquals(users_list_1[3], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[4], User.objects.get(username='username4'))

        # Comprueba que devuelve None si algún usuario del fichero no existe
        file2 = open('census/files/testfiles/testgroup2.txt', 'rb')
        users_list_2 = readTxtFile(file2)
        self.assertEquals(users_list_2, None)


    # Prueba la función "createGroup"
    def test_create_group(self):

        # Comprueba que crea grupo nuevo PRIVADO
        name_1 = 'Grupo 3'
        users_list_1 = User.objects.all()
        res1 = createGroup(name_1, users_list_1, False)
        self.assertTrue(res1)
        
        res2 = True
        try:
            ParentGroup.objects.get(name='Grupo 3', isPublic=False)
        except:
            res2=False

        self.assertTrue(res2)

        # Comprueba que crea grupo nuevo PUBLICO
        name_2 = 'Grupo 4'
        users_list_2 = User.objects.all()
        res3 = createGroup(name_2, users_list_2, True)
        self.assertTrue(res3)
        
        res4 = True
        try:
            ParentGroup.objects.get(name='Grupo 4', isPublic=True)
        except:
            res4=False

        self.assertTrue(res4)


        # Comprueba que actualiza 'Grupo 1'
        name_3 = 'Grupo 1'
        users_list_3 = []
        u1 = User.objects.get(username='username1')
        users_list_3.append(u1)


        res5 = createGroup(name_3, users_list_3, False)
        self.assertFalse(res5)
        self.assertEquals(len(User.objects.filter(groups__name='Grupo 1')), 1)


    # Prueba la función "readExcelFile"
    def test_read_excel_file(self):
        users_list_1 = readExcelFile('census/files/testfiles/testgroup1.xlsx')

        # Comproba cada usuario de la lista devuelta
        self.assertTrue(len(users_list_1)==5)
        self.assertEquals(users_list_1[0], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username5'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))
        self.assertEquals(users_list_1[3], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[4], User.objects.get(username='username4'))

        # Comprueba que devuelve None si algún usuario del fichero no existe
        users_list_2 = readExcelFile('census/files/testfiles/testgroup2.xlsx')
        self.assertEquals(users_list_2, None)


    # Prueba la función "auxUsersList"
    def test_aux_users_list(self):
        username_list_1 = ['username1', 'username2', 'username3']
        users_list_1 = auxUsersList(username_list_1)
        
        self.assertEquals(users_list_1[0], User.objects.get(username='username1'))
        self.assertEquals(users_list_1[1], User.objects.get(username='username2'))
        self.assertEquals(users_list_1[2], User.objects.get(username='username3'))

        username_list_2 = ['username1', 'username2', 'usernamequenoexiste']
        users_list_2 = auxUsersList(username_list_2)

        self.assertEquals(users_list_2, None)


    # Prueba la función "writeInExcelUsernames"
    def test_write_in_excel_usernames(self):
        # Fichero correcto
        users = User.objects.all()
        path = 'census/files/temp_export.xlsx'
        name = 'temp_export.xlsx'
        writeInExcelUsernames(users, path, name)

        # Leer el fichero excel
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        max_row = sheet.max_row
        username_list = []
        for i in range(1, max_row+ 1):
            cell = sheet.cell(row = i, column = 1)
            username_list.append(cell.value)

        # Comprueba cada usuario
        for i in range(0, len(users)):
            self.assertEquals(users[i].username, username_list[i])


class ImportAndExportGroupSeleniumTestCase(SeleniumBaseTestCase):

    def setUp(self):
        g1 = ParentGroup(name='Grupo 1', pk=100)
        g1.save()

        g2 = ParentGroup(name='Grupo 2', pk=101)
        g2.save()

        u1 = User(username='username1', password='password')
        u1.save()
        u1.groups.set([g1, g2])
        u1.save()

        u2 = User(username='username2', password='password')
        u2.save()
        u2.groups.set([g1])
        u2.save()

        u3 = User(username='username3', password='password')
        u3.save()
        u3.groups.set([g1])
        u3.save()

        u4 = User(username='username4', password='password')
        u4.save()
        u4.groups.set([g1])
        u4.save()


        u5 = User(username='username5', password='password')
        u5.save()
        u5.groups.set([g1])
        u5.save()

        return super().setUp()


    def test_import_group(self):
        self.login()
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 3')
        self.driver.find_element_by_id('id_is_public').click()
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/census/files/testfiles/testgroup1.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        result = True
        try:
            self.driver.find_element_by_class_name('success')
        except NoSuchElementException:
            result = False
        self.assertTrue(result)


        self.driver.find_element_by_id('id_name').clear()
        self.driver.find_element_by_id('id_name').send_keys('Grupo 4')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/census/files/testfiles/testgroup1.xlsx")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        result = True
        try:
            self.driver.find_element_by_class_name('success')
        except NoSuchElementException:
            result = False
        self.assertTrue(result)

        # Comprueba que existen 4 grupos (setUp + 2 grupos creados)
        self.driver.get(f"{self.live_server_url}/admin/auth/group/")
        self.assertEquals(len(self.driver.find_elements_by_class_name(name='field-__str__')), 4)


    # Prueba si se introduce un nombre de grupo ya existente
    def test_import_group_name_already_exists(self):
        self.login()
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 2')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/census/files/testfiles/testgroup3.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        # Comprueba que aparece un mensaje de éxito
        self.driver.find_element_by_class_name('success')


        # Comprueba que username1 no pertenece al grupo 2
        self.driver.get(f"{self.live_server_url}/admin/auth/user/")
        self.driver.find_element_by_link_text("username1").click()
        result = True
        try:
            self.driver.find_element_by_xpath("//select[@name='groups']/option[text()='Grupo 2']")
        except NoSuchElementException:
            result = False

        self.assertFalse(result)


        # Comprueba que username2 sí pertenece al grupo 2
        self.driver.get(f"{self.live_server_url}/admin/auth/user/")
        self.driver.find_element_by_link_text("username2").click()
        result = True
        try:
            self.driver.find_element_by_xpath("//select[@name='groups']/option[text()='Grupo 2']")
        except NoSuchElementException:
            result = False

        self.assertTrue(result)


    # Prueba qué ocurre si se envía el formulario con el nombre vacío o solo
    def test_import_group_name_empty(self):
        self.login()
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys(' ')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/census/files/testfiles/testgroup1.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        # Comprueba que se muestra un mensaje de error
        result = True
        try:
            self.driver.find_element_by_class_name('errorlist')
        except NoSuchElementException:
            result = False

        self.assertTrue(result)



    # Prueba si se introduce un fichero con un username que no existe
    def test_import_wrong_username(self):
        self.login()
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.driver.find_element_by_id('id_name').send_keys('Grupo 2')
        self.driver.find_element_by_id('id_file').send_keys(os.getcwd() + "/census/files/testfiles/testgroup2.txt")
        self.driver.find_element_by_xpath("//input[@value='Importar']").click()

        # Comprueba que aparece un mensaje de error
        self.driver.find_element_by_class_name('error')


    # Prueba si se intenta importar sin ser superuser
    def test_import_without_being_superuser(self):
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.assertFalse(re.fullmatch(f'{self.live_server_url}/census/groups/import/', self.driver.current_url))

        self.login(username='username1', password='password')
        self.driver.get(f"{self.live_server_url}/census/groups/import/")
        self.assertFalse(re.fullmatch(f'{self.live_server_url}/census/groups/import/', self.driver.current_url))


    def test_export_group(self):
        self.login()
        self.driver.get(f"{self.live_server_url}/census/groups/export/")
        select = Select(self.driver.find_element_by_id('id_group'))
        select.select_by_visible_text('Grupo 1')
        self.driver.find_element_by_xpath("//input[@value='Exportar']").click()

        # Espero a que termine la descarga
        time.sleep(3)
        # Obtenemos la ruta de descarga, cambia dependiendo del valor de options.headless
        download_file_path = 'export_group.xlsx'
        if (not os.path.exists('export_group.xlsx')):
            download_file_path = os.path.join(os.path.expanduser('~'), 'Descargas/export_group.xlsx') 
        # Compruebo si el fichero existe
        self.assertEquals(os.path.exists(download_file_path), True)

        # Compruebo si todos los usuarios del grupo están
        workbook = openpyxl.load_workbook(download_file_path)
        sheet = workbook.active

        # Lee el excel y almacena en username_list los nombres de usuario
        usernames = ['username1', 'username2', 'username3', 'username4', 'username5']
        max_row = sheet.max_row
        for i in range(1, max_row+ 1):
            cell = sheet.cell(row = i, column = 1)
            self.assertEquals(cell.value in usernames, True)

        # Eliminamos el fichero descargado
        os.remove(download_file_path) 


    # Prueba si se intenta exportar sin ser superuser
    def test_export_without_being_superuser(self):
        self.driver.get(f"{self.live_server_url}/census/groups/export/")
        self.assertFalse(re.fullmatch(f'{self.live_server_url}/census/groups/export/', self.driver.current_url))

        self.login(username='username1', password='password')
        self.driver.get(f"{self.live_server_url}/census/groups/export/")
        self.assertFalse(re.fullmatch(f'{self.live_server_url}/census/groups/export/', self.driver.current_url))
