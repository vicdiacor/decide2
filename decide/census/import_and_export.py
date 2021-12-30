from django.contrib.auth.models import User, Group
from census.models import ParentGroup
import openpyxl


# Dado un fichero txt (en memoria) con un nombre de usuario en cada línea, 
# devuelve una lista de usuarios o None si alguno de los usuarios indicados no existen.
def readTxtFile(file):
    file_str = file.read().decode('utf-8')
    users_list = auxUsersList(list(file_str.split('\n')))
    return users_list


# Dado un nombre y una lista de usuarios, devuelve True si crea un nuevo grupo y False si el grupo ya existe. 
def createGroup(name, user_list, is_public):

    # Comprueba si existe un Group con ese nombre pero no un ParentGroup
    # En dicho caso, borra el Group
    g2 = None
    try:
        g2 = Group.objects.get(name=name)
    except:
        pass

    if (g2 != None):
        try: 
            g = ParentGroup.objects.get(name=name)
        except:
            g2.delete()
    

    g, is_created = ParentGroup.objects.get_or_create(name=name)

    if (not is_created):
        # Eliminamos todos los usuarios de dicho grupo
        for u in User.objects.filter(groups__name=g.name):
            g.user_set.remove(u)
        g.save()

    # Añadimos todos los usuarios de la lista (crear o actualizar)
    for u in user_list:
        g.user_set.add(u)

    g.isPublic = is_public
    g.save()

    return is_created


# Dado el path de un fichero excel, lee la primera columna y retorna una lista de usuarios o None si 
# algún username no existe
def readExcelFile(path):

    # Abre el excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Lee el excel y almacena en username_list los nombres de usuario
    max_row = sheet.max_row
    username_list = []
    for i in range(1, max_row+ 1):
        cell = sheet.cell(row = i, column = 1)
        username_list.append(cell.value)

    users_list = auxUsersList(username_list)
    return users_list


# Dada una lista de usernames (str), devuelve una lista de Users (tipo User) o None si algún
# username no existe
def auxUsersList(username_list):
    users_list=[]
    for ele in username_list:
        try:
            # Saltamos líneas en blanco.
            if (ele.strip()!=''): 
                u = User.objects.get(username=ele.strip())
                users_list.append(u)
        except:
            print('It does not exist a user with username ' + ele.strip()) 
            return None

    return users_list




# Dada un QuerySet de usuarios (de un grupo concreto),
# crea y escribe en un Excel una columna con todos los usernames (una columna, un 
# username por celda)
def writeInExcelUsernames(users, path, name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = name

    for u in users:
        worksheet.append([u.username])

    workbook.save(filename=path)


